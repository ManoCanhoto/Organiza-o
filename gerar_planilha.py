import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Caminhos (Você só precisará mudar o final da pasta quando mudar o mês)
caminho_processados = r #endereço da pasta
caminho_salvar_excel = os.path.join(caminho_processados, #Nome da panilha)
NOME_MES = "MARÇO" # Mude aqui para os próximos meses


def criar_relatorio_com_resumo():
    dados = []
    arquivos_na_pasta = [f for f in os.listdir(caminho_processados) if os.path.isfile(os.path.join(caminho_processados, f))]


    for arquivo in arquivos_na_pasta:
        if arquivo.endswith('.xlsx'): continue
           
        nome_arquivo, extensao = os.path.splitext(arquivo)
        partes = nome_arquivo.split('_')
       
        nome = f"{partes[0]} {partes[1]}" if len(partes) >= 2 else nome_arquivo
        data = partes[2] if len(partes) >= 3 else "Verificar"
       
        dados.append({
            "Funcionário": nome,
            "Data de Emissão": data,
            "Formato": extensao.upper().replace('.', ''),
            "Nome do Arquivo": arquivo
        })


    if not dados:
        print("Nenhum dado encontrado.")
        return


    # 1. Criar os DataFrames
    df_dados = pd.DataFrame(dados)
    df_resumo = pd.DataFrame([{"Mês": NOME_MES, "Total de Atestados": len(dados)}])


    # 2. Salvar com duas abas
    with pd.ExcelWriter(caminho_salvar_excel, engine='openpyxl') as writer:
        df_dados.to_excel(writer, sheet_name='Lista Detalhada', index=False)
        df_resumo.to_excel(writer, sheet_name='Resumo para Chefia', index=False)


    # 3. Formatação Visual
    wb = load_workbook(caminho_salvar_excel)
   
    # Formata ambas as abas
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cor_azul = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        fonte_branca = Font(color="FFFFFF", bold=True)
        borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                cell.border = borda
                if cell.row == 1:
                    cell.fill = cor_azul
                    cell.font = fonte_branca
                    cell.alignment = Alignment(horizontal="center")
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 4


    wb.save(caminho_salvar_excel)
    print(f"✅ Planilha pronta! Total de {len(dados)} atestados em {NOME_MES}.")


if __name__ == "__main__":
    criar_relatorio_com_resumo()
